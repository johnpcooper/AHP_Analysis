# this script is for calculating all the necessary values for AHP analysis


import xlsxwriter
import numpy as np
# import pandas as pd, script doesn't use pandas at this point
from openpyxl import *
import matplotlib.pyplot as plt
from openpyxl.utils import *
# from tkCommonDialog import Dialog # tkinter module used for visual interface
from Tkinter import Tk
from Tkinter import *
import Tkinter, Tkconstants, tkFileDialog
from tkFileDialog import askopenfilename
from tkCommonDialog import Dialog
import tkMessageBox as mb

class GetValues(object):

    def __init__(self):
        mb.showinfo("AHP_Analysis Guide", "Select the file you would like to analyze.")
        # get the file name from the user using tkinter
        Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
        self.filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file
        print(self.filename)
        # create the workbook object using openpyxl
        self.wb = load_workbook(self.filename)
        # create the sheet object using openpyxl
        self.sheet1 = self.wb.get_sheet_by_name('Sheet1')
        # create the max row and column objects
        self.maxrow = self.sheet1.max_row
        self.maxcol = self.sheet1.max_column

        # create the empty T and V lists
        self.T = []
        self.V = []
        # fill the T and V lists with values from the excel sheet. start at
        # 2 bc row 1 is the header. Row 2 is contains first data points
        for i in range(2, self.maxrow-1):
            self.T.append(self.sheet1.cell(row=i,column=1).value)
            self.V.append(self.sheet1.cell(row=i,column=2).value)

    def derive(self): # derive and extract releveant variables from derivative list
        # create a list called dVdT that is the time derivative of voltage at each point
        self.dVdT = np.diff(self.V) / np.diff(self.T)

        # define minimum, maximum, and 0.05*maximum dVdT
        self.mindVdT = np.amin(self.dVdT)
        self.maxdVdT = np.amax(self.dVdT)
        #print'max dVdT = %f' % self.maxdVdT
        #print'0.05*max dVdT = %f' % self.threshVdT


    def thdvdt(self): # extract information regarding threshold dVdT
        self.threshdVdT = float(0.05*self.maxdVdT)
        # np.argmax returns the index of the highest, or in this case, the first value
        # larger than the value following the '>'in an array.
        # rowth = first value higher than threshold dVdT
        self.rowth = np.argmax(self.dVdT>self.threshdVdT)
        self.Tthresh = float(self.T[self.rowth])
        self.Vmthresh = float(self.V[self.rowth])

        #print 'T at threshold dVdT = %f' % float(self.T[self.rowth])
        #print 'threshold dVdT = %f' % float(self.dVdT[self.rowth])
        #return float(self.dVdT[self.rowth])

    def peak(self): # extract information regarding peak Vm

        self.rowpeak = np.argmax(self.V)
        self.Tpeak = self.T[self.rowpeak]
        self.Vpeak = self.V[self.rowpeak]

    def tvsvpoints(self): # create list containing T, V pairs at predefinend T after V peak
        #Tap = T after peak
        self.Tap = [0]

        # The list of T after peak points is created below. For easier adjustment of which
        # points are chosen you could use variables to define dT for each step, how many points
        # you want to look at with that step etc.
        for i in range(0,5,1):
            self.Tap.append(float(0.3+0.05*i))

        for i in range(0,5,1):
            self.Tap.append(float(0.6+0.1*i))

        for i in range(0,8,1):
            self.Tap.append(float(1.5+0.5*i))

        for i in range(0,5,1):
            self.Tap.append(float(6+1*i))

        #print(self.Tap)

    def ahppoints(self): # find the points after peak voltage (where T = 0) that correspond
                         # to desired points in Tap

        # define dT between data points taken by clampfit (~0.004 ms). Reflects a sampling rate of
        # 250 kHZ
        self.dTperCell = 0.004
        # make list that contains number of dT's you must travel per timepoint
        # chosen in tvsvpoints()
        self.stepsatTap = []
        for i in range(0,len(self.Tap),1):
            self.stepsatTap.append(int(self.Tap[i]/self.dTperCell))

        #print self.stepsatTap
        # generate lists of actual T and V after peak V from data
        self.Tapactual = []
        self.Vapactual = []
        for i in range(0,len(self.Tap),1):
            self.Tapactual.append(self.T[self.rowpeak+self.stepsatTap[i]])
            self.Vapactual.append(self.V[self.rowpeak+self.stepsatTap[i]])

        self.ahppairs = zip(self.Tap,self.Vapactual)

        #print(self.Vapactual)

    def writetosheet(self):
        # Create the processed data workbook and add a worksheet using the
        # xlsxwriter package
        # Tell the user to select processed data output directory
        mb.showinfo("AHP_Analysis Guide",
        "Save your output file. Include the file extension .xlsx"
        )
        # ask the user to 'save as' the output file
        root = Tk()
        root.filename = tkFileDialog.asksaveasfilename(
            initialdir = str(self.filename),
            title = "Select file",
            filetypes = (("xlsx files","*.xlsx"),("all files","*.*"))
            )

        newfilename = str(root.filename)
        processedwb = xlsxwriter.Workbook(newfilename)
        processedws = processedwb.add_worksheet()
        # Crate the list that contains relevant refvalues to report
        refvalues = (
            ['Min dVdT', self.mindVdT],
            ['Max dVdT', self.maxdVdT],
            ['0.05*max dVdT (threshold)', self.threshdVdT],
            ['T threshold', self.Tthresh],
            ['Vm threshold', self.Vmthresh],
            ['T peak', self.Tpeak],
            ['V peak', self.Vpeak]
        )

        # define starting row and column
        row = 0
        col = 0
        #iterate over the data and write it out row by row
        for item, value in (refvalues):
            processedws.write(row, col, item)
            processedws.write(row, col + 1, value)
            row += 1

        # Give a title to T, V pairs plot
        processedws.write(0, 6, 'T after peak')
        processedws.write(0, 7, 'Vm')
        # define starting row and column for the T,V pairs
        row = 1
        col = 6
        for T, V in (self.ahppairs):
            processedws.write(row, col, T)
            processedws.write(row, col + 1, V)
            row += 1

        # save and close the workkbook
        processedwb.close()

    def plotahp(self):
        xval = [x[0] for x in self.ahppairs]
        yval = [x[1] for x in self.ahppairs]

        plt.scatter(xval, yval)
        plt.show()
